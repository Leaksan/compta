import pytest
import io
import openpyxl
from datetime import date
from api.index import db, Transaction, Settings

def test_export_includes_custom_fields(pro_client):
    # 1. Setup: Create a custom field
    pro_client.post('/api/settings', json={
        'custom_fields': [{'id': 'cf_test', 'name': 'Test Field', 'type': 'text'}]
    })

    # 2. Setup: Create a transaction with custom data
    pro_client.post('/api/transactions', json={
        'type': 'income',
        'amount': 1000,
        'category': 'Vente',
        'label': 'Test Custom',
        'date': '2024-01-01',
        'custom_data': {'cf_test': 'Custom Value'}
    })

    # 3. Export
    response = pro_client.get('/api/export?year=2024&month=1')
    assert response.status_code == 200

    buf = io.BytesIO(response.data)
    wb = openpyxl.load_workbook(buf)
    ws = wb.active

    # 4. Verify headers
    headers = [c.value for c in ws[1]]
    assert 'Test Field' in headers

    # 5. Verify data
    custom_col_idx = headers.index('Test Field') + 1
    assert ws.cell(row=2, column=custom_col_idx).value == 'Custom Value'

def test_import_skips_total_row(pro_client):
    # 1. Create an excel file with a TOTAL row
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(['Date', 'Catégorie', 'Libellé', 'Entrée', 'Dépense', 'Quantité'])
    ws.append(['01/01/2024', 'Vente', 'Real TX', 100, 0, 1])
    ws.append(['TOTAL', '', '', 100, 0, ''])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    # 2. Import
    response = pro_client.post('/api/import', data={
        'file': (buf, 'test.xlsx')
    }, content_type='multipart/form-data')

    assert response.status_code == 200
    # Should only import 1 transaction, not 2
    assert '1 transaction(s) importée(s)' in response.json['message']

def test_import_handles_custom_fields(pro_client):
    # 1. Setup: Create a custom field
    pro_client.post('/api/settings', json={
        'custom_fields': [{'id': 'cf_test', 'name': 'Test Field', 'type': 'text'}]
    })

    # 2. Create excel with custom field column
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(['Date', 'Catégorie', 'Libellé', 'Entrée', 'Dépense', 'Quantité', 'Test Field'])
    ws.append(['01/01/2024', 'Vente', 'Custom TX', 100, 0, 1, 'My Custom Value'])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    # 3. Import
    response = pro_client.post('/api/import', data={
        'file': (buf, 'test.xlsx')
    }, content_type='multipart/form-data')

    assert response.status_code == 200

    # 4. Verify transaction in DB
    with pro_client.application.app_context():
        tx = Transaction.query.filter_by(label='Custom TX').first()
        assert tx is not None
        import json
        custom_data = json.loads(tx.custom_data)
        assert custom_data.get('cf_test') == 'My Custom Value'

def test_import_multi_sheet(pro_client):
    # 1. Create a multi-sheet excel file
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Janvier"
    ws1.append(['Date', 'Catégorie', 'Libellé', 'Entrée', 'Dépense', 'Quantité'])
    ws1.append(['01/01/2024', 'Vente', 'TX1', 100, 0, 1])

    ws2 = wb.create_sheet("Février")
    ws2.append(['Date', 'Catégorie', 'Libellé', 'Entrée', 'Dépense', 'Quantité'])
    ws2.append(['01/02/2024', 'Vente', 'TX2', 200, 0, 1])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    # 2. Import
    response = pro_client.post('/api/import', data={
        'file': (buf, 'test.xlsx')
    }, content_type='multipart/form-data')

    assert response.status_code == 200
    assert '2 transaction(s) importée(s)' in response.json['message']

def test_import_robust_headers(pro_client):
    # 1. Create excel with accented and mixed case headers
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(['DATE', 'CATÉGORIE', 'LIBELLÉ', 'ENTRÉE', 'DÉPENSE', 'QUANTITÉ'])
    ws.append(['01/01/2024', 'Vente', 'Accented TX', 100, 0, 1])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    # 2. Import
    response = pro_client.post('/api/import', data={
        'file': (buf, 'test.xlsx')
    }, content_type='multipart/form-data')

    assert response.status_code == 200
    assert '1 transaction(s) importée(s)' in response.json['message']

    # 3. Verify
    with pro_client.application.app_context():
        tx = Transaction.query.filter_by(label='Accented TX').first()
        assert tx is not None
        assert tx.category == 'Vente'
        assert tx.amount == 100
