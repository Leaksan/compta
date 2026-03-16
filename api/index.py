import re
from datetime import timezone
import os
import uuid
import json
import io
import urllib.parse
from datetime import datetime, date
from dateutil.relativedelta import relativedelta
from flask import (
    Flask,
    render_template,
    request,
    redirect,
    url_for,
    jsonify,
    send_file,
    abort
)
from flask_sqlalchemy import SQLAlchemy
from flask_login import (
    LoginManager,
    UserMixin,
    login_user,
    logout_user,
    login_required,
    current_user,
)
from werkzeug.security import generate_password_hash, check_password_hash
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# ─── App Setup ────────────────────────────────────────────────────────────────
app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "compta-secret-key-change-in-prod")
app.config["SQLALCHEMY_DATABASE_URI"] = os.environ.get(
    "DATABASE_URL",
    "postgresql://neondb_owner:npg_viUgEYWc2JD5@ep-red-leaf-anz1mvq7-pooler.c-6.us-east-1.aws.neon.tech/neondb?sslmode=require&channel_binding=require"
)
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False

db = SQLAlchemy(app)
login_manager = LoginManager(app)
login_manager.login_view = "auth"
login_manager.login_message = None

# ─── Constants ────────────────────────────────────────────────────────────────
FREE_PLAN_LIMIT = 20
PREDEFINED_CATEGORIES = {
    "income": ["Vente", "Prestation", "Salaire", "Remboursement", "Autre entrée"],
    "expense": [
        "Loyer",
        "Fournitures",
        "Transport",
        "Salaires versés",
        "Communication",
        "Fiscalité",
        "Autre dépense",
    ],
}
DEFAULT_FIELD_ORDER = [
    "date",
    "category",
    "label",
    "quantity",
    "observation",
    "income",
    "expense",
]
DEFAULT_HIDDEN_FIELDS = ["quantity"]
DEFAULT_FIELD_LABELS = {
    "date": "Date",
    "category": "Catégorie",
    "label": "Libellé",
    "observation": "Observation",
    "income": "Entrée",
    "expense": "Dépense",
    "quantity": "Quantité",
}


# ─── Models ───────────────────────────────────────────────────────────────────
class User(UserMixin, db.Model):
    __tablename__ = "users"
    id = db.Column(db.Integer, primary_key=True)
    first_name = db.Column(db.String(100), nullable=False)
    company_name = db.Column(db.String(200), nullable=False)
    whatsapp = db.Column(db.String(30), unique=True, nullable=False)
    email = db.Column(db.String(200), nullable=True)
    password_hash = db.Column(db.String(256), nullable=False)
    is_pro = db.Column(db.Boolean, default=False)
    plan = db.Column(db.String(20), default="free")
    activated_at = db.Column(db.DateTime, nullable=True)
    created_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc))

    # Relationships
    transactions = db.relationship(
        "Transaction", backref="user", lazy=True, cascade="all, delete-orphan"
    )
    settings = db.relationship(
        "Settings", backref="user", uselist=False, cascade="all, delete-orphan"
    )
    exports = db.relationship(
        "Export", backref="user", lazy=True, cascade="all, delete-orphan"
    )

    def set_password(self, pw):
        self.password_hash = generate_password_hash(pw)

    def check_password(self, pw):
        return check_password_hash(self.password_hash, pw)


class Transaction(db.Model):
    __tablename__ = "transactions"
    id = db.Column(db.String(36), primary_key=True, default=lambda: str(uuid.uuid4()))
    user_id = db.Column(db.Integer, db.ForeignKey("users.id"), nullable=False)
    type = db.Column(db.String(10), nullable=False)  # 'income' | 'expense'
    amount = db.Column(db.Float, nullable=False, default=0)
    quantity = db.Column(db.Integer, default=1)
    label = db.Column(db.String(300), default="")
    category = db.Column(db.String(100), nullable=False)
    observation = db.Column(db.Text, default="")
    date = db.Column(db.Date, nullable=False)
    created_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc))
    custom_data = db.Column(db.Text, default="{}")  # JSON

    def to_dict(self):
        return {
            "id": self.id,
            "type": self.type,
            "amount": self.amount,
            "quantity": self.quantity,
            "label": self.label,
            "category": self.category,
            "observation": self.observation,
            "date": self.date.isoformat(),
            "created_at": self.created_at.isoformat(),
            "custom_data": json.loads(self.custom_data or "{}"),
        }


class Export(db.Model):
    __tablename__ = "exports"
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey("users.id"), nullable=False)
    filename = db.Column(db.String(255), nullable=False)
    year = db.Column(db.Integer, nullable=False)
    month = db.Column(db.Integer, nullable=False)
    end_year = db.Column(db.Integer, nullable=False)
    end_month = db.Column(db.Integer, nullable=False)
    created_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc))


class Settings(db.Model):
    __tablename__ = "settings"
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(
        db.Integer, db.ForeignKey("users.id"), nullable=False, unique=True
    )
    currency = db.Column(db.String(10), default="FCFA")
    categories_json = db.Column(db.Text, default=json.dumps(PREDEFINED_CATEGORIES))
    custom_categories_json = db.Column(db.Text, default="[]")
    custom_fields_json = db.Column(db.Text, default="[]")
    field_order_json = db.Column(db.Text, default=json.dumps(DEFAULT_FIELD_ORDER))
    hidden_fields_json = db.Column(db.Text, default=json.dumps(DEFAULT_HIDDEN_FIELDS))
    field_labels_json = db.Column(db.Text, default=json.dumps(DEFAULT_FIELD_LABELS))

    @property
    def categories(self):
        return json.loads(self.categories_json or json.dumps(PREDEFINED_CATEGORIES))

    @categories.setter
    def categories(self, v):
        self.categories_json = json.dumps(v)

    @property
    def custom_fields(self):
        return json.loads(self.custom_fields_json or "[]")

    @custom_fields.setter
    def custom_fields(self, v):
        self.custom_fields_json = json.dumps(v)

    @property
    def field_order(self):
        return json.loads(self.field_order_json or json.dumps(DEFAULT_FIELD_ORDER))

    @field_order.setter
    def field_order(self, v):
        self.field_order_json = json.dumps(v)

    @property
    def hidden_fields(self):
        return json.loads(self.hidden_fields_json or json.dumps(DEFAULT_HIDDEN_FIELDS))

    @hidden_fields.setter
    def hidden_fields(self, v):
        self.hidden_fields_json = json.dumps(v)

    @property
    def field_labels(self):
        return json.loads(self.field_labels_json or json.dumps(DEFAULT_FIELD_LABELS))

    @field_labels.setter
    def field_labels(self, v):
        self.field_labels_json = json.dumps(v)

    def to_dict(self):
        return {
            "currency": self.currency,
            "categories": self.categories,
            "custom_fields": self.custom_fields,
            "field_order": self.field_order,
            "hidden_fields": self.hidden_fields,
            "field_labels": self.field_labels,
        }


class GlobalSettings(db.Model):
    __tablename__ = "global_settings"
    id = db.Column(db.Integer, primary_key=True)
    whatsapp_number = db.Column(db.String(20), default="")
    admin_key = db.Column(db.String(50), default="")


@login_manager.user_loader
def load_user(uid):
    return db.session.get(User, int(uid))


# ─── Helpers ──────────────────────────────────────────────────────────────────
def get_or_create_settings(user):
    if not user.settings:
        s = Settings(user_id=user.id)
        db.session.add(s)
        db.session.commit()
    return user.settings


def format_currency(amount, currency):
    """Format amount with currency for display"""
    try:
        amount = float(amount)
        if currency == "FCFA":
            return f"{int(amount):,}".replace(",", " ") + " FCFA"
        elif currency == "EUR":
            return f"{amount:,.0f} €".replace(",", " ")
        elif currency == "USD":
            return f"${amount:,.0f}".replace(",", " ")
        else:
            return f"{amount:,.0f} {currency}".replace(",", " ")
    except:
        return f"0 {currency}"


def get_transactions_for_period(user_id, start_date, end_date):
    return (
        Transaction.query.filter_by(user_id=user_id)
        .filter(Transaction.date >= start_date, Transaction.date <= end_date)
        .order_by(Transaction.date.desc(), Transaction.created_at.desc())
        .all()
    )


def get_available_months(user_id):
    """Returns list of (year, month) tuples that have transactions, sorted desc"""
    from sqlalchemy import extract

    rows = (
        db.session.query(
            extract("year", Transaction.date).label("year"),
            extract("month", Transaction.date).label("month"),
        )
        .filter(Transaction.user_id == user_id)
        .distinct()
        .order_by(db.text("year DESC, month DESC"))
        .all()
    )
    return [(int(r.year), int(r.month)) for r in rows]


def count_transactions_this_month(user_id, year, month):
    from sqlalchemy import extract

    return (
        Transaction.query.filter_by(user_id=user_id)
        .filter(
            extract("year", Transaction.date) == year,
            extract("month", Transaction.date) == month,
        )
        .count()
    )


def get_performance_observation(txs, prev_txs, period_start, period_end, fmt_fn):
    if not txs and not prev_txs:
        return None
    if not txs:
        return {
            "text": "Nouveau mois, nouveau départ ! Ajoutez vos premières transactions.",
            "type": "neutral",
        }

    today = date.today()
    is_current_month = (
        period_start.year == today.year and period_start.month == today.month
    )
    show_balance_comparison = (not is_current_month) or today.day > 21

    total_income = sum(t.amount for t in txs if t.type == "income")
    total_expense = sum(t.amount for t in txs if t.type == "expense")
    net_balance = total_income - total_expense

    prev_income = sum(t.amount for t in prev_txs if t.type == "income")
    prev_expense = sum(t.amount for t in prev_txs if t.type == "expense")
    prev_net = prev_income - prev_expense

    is_same_month = (
        period_start.year == period_end.year and period_start.month == period_end.month
    )
    prev_month_name = (period_start - relativedelta(months=1)).strftime("%B")
    prev_period_text = (
        f"au mois de {prev_month_name}" if is_same_month else "à la période précédente"
    )

    def cat_totals(ts, t_type):
        d = {}
        for t in ts:
            if t.type == t_type:
                d[t.category] = d.get(t.category, 0) + t.amount
        return d

    cur_inc = cat_totals(txs, "income")
    prev_inc = cat_totals(prev_txs, "income")
    cur_exp = cat_totals(txs, "expense")
    prev_exp = cat_totals(prev_txs, "expense")

    insights = []

    if show_balance_comparison:
        if not prev_txs:
            if net_balance > 0:
                insights.append(
                    {
                        "text": f"C'est un excellent mois ! Vous avez réalisé {fmt_fn(net_balance)} de bénéfice.",
                        "type": "positive",
                        "score": 100,
                    }
                )
            elif net_balance < 0:
                insights.append(
                    {
                        "text": f"Attention, vos dépenses ont dépassé vos revenus de {fmt_fn(abs(net_balance))}.",
                        "type": "negative",
                        "score": 100,
                    }
                )
        else:
            diff = net_balance - prev_net
            if diff > 0:
                insights.append(
                    {
                        "text": f"Bravo ! Votre solde s'est amélioré de {fmt_fn(diff)} par rapport {prev_period_text}.",
                        "type": "positive",
                        "score": 100,
                    }
                )
            elif diff < 0:
                insights.append(
                    {
                        "text": f"Attention, votre solde a baissé de {fmt_fn(abs(diff))} par rapport {prev_period_text}.",
                        "type": "negative",
                        "score": 100,
                    }
                )
            else:
                insights.append(
                    {
                        "text": f"Votre solde est stable par rapport {prev_period_text}.",
                        "type": "neutral",
                        "score": 100,
                    }
                )

    for cat, cur in cur_inc.items():
        prev = prev_inc.get(cat, 0)
        if prev > 0 and cur > prev:
            growth = cur - prev
            pct = (growth / prev) * 100
            if pct > 50:
                insights.append(
                    {
                        "text": f'Excellente nouvelle ! Vos revenus en "{cat}" ont explosé (+{fmt_fn(growth)}).',
                        "type": "positive",
                        "score": 90,
                    }
                )
            else:
                insights.append(
                    {
                        "text": f'Belle progression ! Vos revenus en "{cat}" ont augmenté de {fmt_fn(growth)}.',
                        "type": "positive",
                        "score": 70,
                    }
                )

    for cat, prev in prev_exp.items():
        cur = cur_exp.get(cat, 0)
        if prev > 0 and cur < prev:
            reduction = prev - cur
            pct = (reduction / prev) * 100
            if pct > 20 and cur > 0:
                insights.append(
                    {
                        "text": f'Super effort ! Vos dépenses en "{cat}" ont baissé de {fmt_fn(reduction)}.',
                        "type": "positive",
                        "score": 80,
                    }
                )
            elif cur == 0:
                insights.append(
                    {
                        "text": f'Parfait ! Aucune dépense en "{cat}" ce mois-ci, contre {fmt_fn(prev)} {prev_period_text}.',
                        "type": "positive",
                        "score": 85,
                    }
                )

    for cat, cur in cur_exp.items():
        prev = prev_exp.get(cat, 0)
        if prev > 0 and cur > prev:
            spike = cur - prev
            pct = (spike / prev) * 100
            if pct > 30:
                insights.append(
                    {
                        "text": f'Attention, vos dépenses en "{cat}" ont fortement augmenté (+{fmt_fn(spike)}).',
                        "type": "negative",
                        "score": 75,
                    }
                )

    if not total_expense and total_income > 0:
        insights.append(
            {
                "text": "Zéro dépense enregistrée pour le moment, belle épargne en perspective !",
                "type": "positive",
                "score": 60,
            }
        )

    if cur_inc:
        top = max(cur_inc.items(), key=lambda x: x[1])
        insights.append(
            {
                "text": f'La catégorie "{top[0]}" est votre meilleure source de revenus ({fmt_fn(top[1])}).',
                "type": "positive",
                "score": 50,
            }
        )

    if cur_exp:
        top = max(cur_exp.items(), key=lambda x: x[1])
        insights.append(
            {
                "text": f'Votre principale dépense ce mois-ci est "{top[0]}" ({fmt_fn(top[1])}).',
                "type": "neutral",
                "score": 40,
            }
        )

    if not insights:
        return {
            "text": "Commencez à ajouter des transactions pour voir votre analyse.",
            "type": "neutral",
        }

    insights.sort(key=lambda x: -x["score"])
    top_score = insights[0]["score"]
    top_insights = [i for i in insights if i["score"] >= top_score - 20]
    idx = today.day % len(top_insights)
    chosen = top_insights[idx]
    return {"text": chosen["text"], "type": chosen["type"]}


# ─── Auth Routes ──────────────────────────────────────────────────────────────
@app.route("/auth", methods=["GET", "POST"])
def auth():
    if current_user.is_authenticated:
        return redirect(url_for("dashboard"))
    return render_template("auth.html")


@app.route("/api/auth/register", methods=["POST"])
def api_register():
    data = request.get_json()
    whatsapp = (data.get("countryCode", "") + data.get("whatsapp", "")).strip()
    if not whatsapp:
        whatsapp = data.get("fullWhatsapp", "")
    if User.query.filter_by(whatsapp=whatsapp).first():
        return jsonify(
            {"success": False, "error": "Ce numéro WhatsApp est déjà utilisé."}
        )
    user = User(
        first_name=data.get("firstName", ""),
        company_name=data.get("companyName", ""),
        whatsapp=whatsapp,
        email=data.get("email") or None,
    )
    user.set_password(data.get("password", ""))
    db.session.add(user)
    db.session.commit()
    s = Settings(user_id=user.id)
    db.session.add(s)
    db.session.commit()
    login_user(user, remember=True)
    return jsonify(
        {
            "success": True,
            "user": {
                "id": user.id,
                "first_name": user.first_name,
                "company_name": user.company_name,
                "whatsapp": user.whatsapp,
                "email": user.email,
                "is_pro": user.is_pro,
                "plan": user.plan,
            },
        }
    )


@app.route("/api/auth/login", methods=["POST"])
def api_login():
    data = request.get_json()
    whatsapp = (data.get("countryCode", "") + data.get("whatsapp", "")).strip()
    if not whatsapp:
        whatsapp = data.get("fullWhatsapp", "")
    password = data.get("password", "")
    user = User.query.filter_by(whatsapp=whatsapp).first()
    if not user or not user.check_password(password):
        return jsonify(
            {"success": False, "error": "Numéro WhatsApp ou mot de passe incorrect."}
        )
    login_user(user, remember=True)
    return jsonify(
        {
            "success": True,
            "user": {
                "id": user.id,
                "first_name": user.first_name,
                "company_name": user.company_name,
                "whatsapp": user.whatsapp,
                "email": user.email,
                "is_pro": user.is_pro,
                "plan": user.plan,
                "activated_at": user.activated_at.isoformat()
                if user.activated_at
                else None,
            },
        }
    )


@app.route("/api/auth/logout", methods=["POST"])
@login_required
def api_logout():
    logout_user()
    return jsonify({"success": True})


@app.route("/logout")
@login_required
def logout():
    logout_user()
    return redirect(url_for("auth"))


# ─── Dashboard Route ──────────────────────────────────────────────────────────
@app.route("/")
@login_required
def dashboard():
    settings = get_or_create_settings(current_user)

    # Parse period from query params
    today = date.today()
    year = request.args.get("year", today.year, type=int)
    month = request.args.get("month", today.month, type=int)
    end_year = request.args.get("end_year", year, type=int)
    end_month = request.args.get("end_month", month, type=int)

    period_start = date(year, month, 1)
    period_end_raw = date(end_year, end_month, 1)
    # Last day of end month
    import calendar

    period_end = date(end_year, end_month, calendar.monthrange(end_year, end_month)[1])

    # Previous period
    months_diff = (end_year - year) * 12 + (end_month - month) + 1
    prev_start = period_start - relativedelta(months=months_diff)
    prev_end_raw = period_end_raw - relativedelta(months=months_diff)
    prev_end = date(
        prev_end_raw.year,
        prev_end_raw.month,
        calendar.monthrange(prev_end_raw.year, prev_end_raw.month)[1],
    )

    txs = get_transactions_for_period(current_user.id, period_start, period_end)
    prev_txs = get_transactions_for_period(current_user.id, prev_start, prev_end)

    total_income = sum(t.amount for t in txs if t.type == "income")
    total_expense = sum(t.amount for t in txs if t.type == "expense")
    net_balance = total_income - total_expense

    currency = settings.currency
    fmt = lambda x: format_currency(x, currency)

    # Check if free plan limit reached for this month
    month_count = count_transactions_this_month(current_user.id, year, month)
    can_add = current_user.is_pro or (month_count < FREE_PLAN_LIMIT)
    remaining = (
        max(0, FREE_PLAN_LIMIT - month_count) if not current_user.is_pro else None
    )

    observation = get_performance_observation(
        txs, prev_txs, period_start, period_end, fmt
    )

    # Period name

    MONTHS_FR = [
        "janvier",
        "février",
        "mars",
        "avril",
        "mai",
        "juin",
        "juillet",
        "août",
        "septembre",
        "octobre",
        "novembre",
        "décembre",
    ]
    start_str = f"{MONTHS_FR[period_start.month - 1]} {period_start.year}"
    end_str = f"{MONTHS_FR[period_end_raw.month - 1]} {period_end_raw.year}"
    period_name = start_str if start_str == end_str else f"{start_str} – {end_str}"

    # Nav months
    prev_month_date = period_start - relativedelta(months=1)
    next_month_date = period_start + relativedelta(months=1)

    # Greeting
    hour = datetime.now().hour
    greeting = (
        f"Bonjour {current_user.first_name}"
        if hour < 18
        else f"Bonsoir {current_user.first_name}"
    )

    # Chart data
    chart_data = build_chart_data(txs, currency)

    return render_template(
        "dashboard.html",
        transactions=txs,
        total_income=fmt(total_income),
        total_expense=fmt(total_expense),
        net_balance=fmt(net_balance),
        net_balance_raw=net_balance,
        period_name=period_name,
        year=year,
        month=month,
        end_year=end_year,
        end_month=end_month,
        prev_year=prev_month_date.year,
        prev_month=prev_month_date.month,
        next_year=next_month_date.year,
        next_month=next_month_date.month,
        settings=settings,
        user=current_user,
        can_add=can_add,
        remaining=remaining,
        observation=observation,
        greeting=greeting,
        chart_data=json.dumps(chart_data),
        fmt=fmt,
        currency=currency,
        MONTHS_FR=MONTHS_FR,
        is_custom_period=(year != end_year or month != end_month),
    )


def build_chart_data(txs, currency):
    groups = {}
    for t in sorted(txs, key=lambda x: x.date):
        key = t.date.strftime("%d %b")
        ts = t.date.isoformat()
        if key not in groups:
            groups[key] = {"date": key, "income": 0, "expense": 0, "ts": ts}
        if t.type == "income":
            groups[key]["income"] += t.amount
        else:
            groups[key]["expense"] += t.amount
    days = sorted(groups.values(), key=lambda x: x["ts"])
    running = 0
    result = []
    for d in days:
        running += d["income"] - d["expense"]
        result.append(
            {
                "date": d["date"],
                "income": d["income"],
                "expense": d["expense"],
                "balance": running,
            }
        )
    return result


# ─── History Route ────────────────────────────────────────────────────────────
@app.route("/history")
@login_required
def history():
    settings = get_or_create_settings(current_user)
    months = get_available_months(current_user.id)
    month_data = []
    MONTHS_FR = [
        "janvier",
        "février",
        "mars",
        "avril",
        "mai",
        "juin",
        "juillet",
        "août",
        "septembre",
        "octobre",
        "novembre",
        "décembre",
    ]
    import calendar

    fmt = lambda x: format_currency(x, settings.currency)
    for y, m in months:
        start = date(y, m, 1)
        end = date(y, m, calendar.monthrange(y, m)[1])
        txs = get_transactions_for_period(current_user.id, start, end)
        ti = sum(t.amount for t in txs if t.type == "income")
        te = sum(t.amount for t in txs if t.type == "expense")
        net = ti - te
        month_data.append(
            {
                "year": y,
                "month": m,
                "name": f"{MONTHS_FR[m - 1].capitalize()} {y}",
                "income": fmt(ti),
                "expense": fmt(te),
                "net": fmt(net),
                "net_raw": net,
                "count": len(txs),
            }
        )
    exports = Export.query.filter_by(user_id=current_user.id).order_by(Export.created_at.desc()).all()
    return render_template(
        "history.html", months=month_data, exports=exports, settings=settings, user=current_user
    )


# ─── Settings Route ───────────────────────────────────────────────────────────
@app.route("/settings", methods=["GET"])
@login_required
def settings_view():
    settings = get_or_create_settings(current_user)
    end_date = None
    if current_user.is_pro and current_user.plan and current_user.activated_at:
        if current_user.plan == "monthly":
            end = current_user.activated_at + relativedelta(months=1)
            end_date = end.strftime("%d/%m/%Y")
        elif current_user.plan == "quarterly":
            end = current_user.activated_at + relativedelta(months=3)
            end_date = end.strftime("%d/%m/%Y")
        elif current_user.plan in ("yearly", "annual"):
            end = current_user.activated_at + relativedelta(years=1)
            end_date = end.strftime("%d/%m/%Y")
    return render_template(
        "settings.html", settings=settings, user=current_user, end_date=end_date
    )


@app.route("/api/settings", methods=["POST"])
@login_required
def api_save_settings():
    data = request.get_json()
    s = get_or_create_settings(current_user)
    if "currency" in data:
        s.currency = data["currency"]
    if "field_labels" in data:
        s.field_labels = data["field_labels"]
    if "categories" in data:
        s.categories = data["categories"]
    if "custom_fields" in data:
        s.custom_fields = data["custom_fields"]
    if "field_order" in data:
        s.field_order = data["field_order"]
    if "hidden_fields" in data:
        s.hidden_fields = data["hidden_fields"]
    db.session.commit()
    return jsonify({"success": True})


@app.route("/api/profile", methods=["POST"])
@login_required
def api_save_profile():
    data = request.get_json()
    if data.get("firstName"):
        current_user.first_name = data["firstName"]
    if data.get("companyName"):
        current_user.company_name = data["companyName"]
    if data.get("email") is not None:
        current_user.email = data["email"] or None
    if data.get("password"):
        current_user.set_password(data["password"])
    db.session.commit()
    return jsonify({"success": True})


# ─── Pro Route ────────────────────────────────────────────────────────────────
@app.route("/pro")
@login_required
def pro_view():
    settings = get_or_create_settings(current_user)
    return render_template("pro.html", settings=settings, user=current_user)


@app.route("/api/activate-pro", methods=["POST"])
@login_required
def api_activate_pro():
    data = request.get_json()
    code = data.get("code", "").strip().upper()
    if code == "PROTEST":
        current_user.is_pro = True
        current_user.plan = "pro"
        current_user.activated_at = datetime.now(timezone.utc)
        db.session.commit()
        return jsonify({"success": True})
    return jsonify({"success": False, "error": "Code promo invalide."})


@app.route("/api/subscribe-pro", methods=["POST"])
@login_required
def api_subscribe_pro():
    data = request.get_json()
    plan = data.get("plan")
    if plan not in ["monthly", "quarterly", "annual"]:
        return jsonify({"success": False, "error": "Plan invalide."})
    plan_names = {"monthly": "Mensuel", "quarterly": "Trimestriel", "annual": "Annuel"}
    message = f"Bonjour, je souhaite m'abonner au plan PRO {plan_names[plan]}. Nom: {current_user.first_name or ''}, Entreprise: {current_user.company_name or ''}, WhatsApp: {current_user.whatsapp or ''}, Email: {current_user.email or ''}"
    encoded_message = urllib.parse.quote(message)

    gs = GlobalSettings.query.first()
    wa_number = gs.whatsapp_number if gs and gs.whatsapp_number else "24177000000"
    if not wa_number.startswith("+"):
        wa_number = "+" + wa_number
    whatsapp_url = f"https://wa.me/{wa_number}?text={encoded_message}"
    return jsonify({"success": True, "whatsapp_url": whatsapp_url})


# ─── Transaction API ──────────────────────────────────────────────────────────
@app.route("/api/transactions", methods=["GET"])
@login_required
def api_get_transactions():
    year = request.args.get("year", date.today().year, type=int)
    month = request.args.get("month", date.today().month, type=int)
    import calendar

    start = date(year, month, 1)
    end = date(year, month, calendar.monthrange(year, month)[1])
    txs = get_transactions_for_period(current_user.id, start, end)
    return jsonify([t.to_dict() for t in txs])


@app.route("/api/transactions", methods=["POST"])
@login_required
def api_create_transaction():
    data = request.get_json()
    year = int(data.get("date", date.today().isoformat())[:4])
    month = int(data.get("date", date.today().isoformat())[5:7])
    # Check free plan limit
    if not current_user.is_pro:
        count = count_transactions_this_month(current_user.id, year, month)
        if count >= FREE_PLAN_LIMIT:
            return jsonify(
                {
                    "success": False,
                    "error": f"Limite de {FREE_PLAN_LIMIT} transactions atteinte. Passez en PRO !",
                }
            )
    try:
        tx_date = date.fromisoformat(data["date"])
    except:
        tx_date = date.today()
    t = Transaction(
        id=str(uuid.uuid4()),
        user_id=current_user.id,
        type=data.get("type", "expense"),
        amount=float(data.get("amount", 0)),
        quantity=int(data.get("quantity", 1)),
        label=data.get("label", ""),
        category=data.get("category", ""),
        observation=data.get("observation", ""),
        date=tx_date,
        custom_data=json.dumps(data.get("custom_data", {})),
    )
    db.session.add(t)
    db.session.commit()
    return jsonify({"success": True, "transaction": t.to_dict()})


@app.route("/api/transactions/<tx_id>", methods=["PUT"])
@login_required
def api_update_transaction(tx_id):
    t = Transaction.query.filter_by(id=tx_id, user_id=current_user.id).first_or_404()
    data = request.get_json()
    if "type" in data:
        t.type = data["type"]
    if "amount" in data:
        t.amount = float(data["amount"])
    if "quantity" in data:
        t.quantity = int(data["quantity"])
    if "label" in data:
        t.label = data["label"]
    if "category" in data:
        t.category = data["category"]
    if "observation" in data:
        t.observation = data["observation"]
    if "date" in data:
        try:
            t.date = date.fromisoformat(data["date"])
        except:
            pass
    if "custom_data" in data:
        t.custom_data = json.dumps(data["custom_data"])
    db.session.commit()
    return jsonify({"success": True, "transaction": t.to_dict()})


@app.route("/api/transactions/<tx_id>", methods=["DELETE"])
@login_required
def api_delete_transaction(tx_id):
    t = Transaction.query.filter_by(id=tx_id, user_id=current_user.id).first_or_404()
    db.session.delete(t)
    db.session.commit()
    return jsonify({"success": True})


# ─── Export Excel ─────────────────────────────────────────────────────────────
@app.route("/api/export")
@login_required
def api_export():
    year = request.args.get("year", date.today().year, type=int)
    month = request.args.get("month", date.today().month, type=int)
    end_year = request.args.get("end_year", year, type=int)
    end_month = request.args.get("end_month", month, type=int)
    import calendar

    start = date(year, month, 1)
    end = date(end_year, end_month, calendar.monthrange(end_year, end_month)[1])

    settings = get_or_create_settings(current_user)
    labels = settings.field_labels

    MONTHS_FR = [
        "janvier", "février", "mars", "avril", "mai", "juin",
        "juillet", "août", "septembre", "octobre", "novembre", "décembre",
    ]

    wb = openpyxl.Workbook()
    # Remove default sheet
    if wb.active:
        wb.remove(wb.active)

    # Calculate month ranges
    month_ranges = []
    curr = start
    while curr <= end:
        m_start = curr
        m_end = date(curr.year, curr.month, calendar.monthrange(curr.year, curr.month)[1])
        month_ranges.append((m_start, m_end))
        curr += relativedelta(months=1)

    green = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
    red = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    header_fill = PatternFill(start_color="171717", end_color="171717", fill_type="solid")

    for m_start, m_end in month_ranges:
        sheet_title = f"{MONTHS_FR[m_start.month - 1]} {m_start.year}"
        ws = wb.create_sheet(title=sheet_title)

        # Header row
        headers = [
            labels.get("date", "Date"),
            labels.get("category", "Catégorie"),
            labels.get("label", "Libellé"),
            labels.get("observation", "Observation"),
            labels.get("quantity", "Quantité"),
            labels.get("income", "Entrée"),
            labels.get("expense", "Dépense"),
            "Solde cumulé",
        ]
        for i, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=i, value=h)
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = header_fill
            c.alignment = Alignment(horizontal="center")

        txs = get_transactions_for_period(current_user.id, m_start, m_end)
        txs_sorted = sorted(txs, key=lambda x: x.date)

        running = 0
        for row_idx, t in enumerate(txs_sorted, 2):
            running += t.amount if t.type == "income" else -t.amount
            ws.cell(row=row_idx, column=1, value=t.date.strftime("%d/%m/%Y"))
            ws.cell(row=row_idx, column=2, value=t.category)
            ws.cell(row=row_idx, column=3, value=t.label)
            ws.cell(row=row_idx, column=4, value=t.observation)
            ws.cell(row=row_idx, column=5, value=t.quantity)

            inc_val = t.amount if t.type == "income" else 0
            exp_val = t.amount if t.type == "expense" else 0

            inc_cell = ws.cell(row=row_idx, column=6, value=inc_val)
            exp_cell = ws.cell(row=row_idx, column=7, value=exp_val)
            bal_cell = ws.cell(row=row_idx, column=8, value=running)

            if t.type == "income":
                inc_cell.fill = green
            elif t.type == "expense":
                exp_cell.fill = red

            bal_cell.fill = green if running >= 0 else red

        # Totals row
        total_row = len(txs_sorted) + 2
        ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
        total_inc = sum(t.amount for t in txs_sorted if t.type == "income")
        total_exp = sum(t.amount for t in txs_sorted if t.type == "expense")

        c_inc = ws.cell(row=total_row, column=6, value=total_inc)
        c_inc.font = Font(bold=True, color="059669")
        c_exp = ws.cell(row=total_row, column=7, value=total_exp)
        c_exp.font = Font(bold=True, color="DC2626")

        # Auto-width
        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    comp_name = re.sub(r'[^a-zA-Z0-9]', '_', current_user.company_name or 'Export')
    filename = f"{comp_name}_du{start.strftime('%d_%m')}_au_{end.strftime('%d_%m')}.xlsx"
    # Record export in history
    record_export = request.args.get("record", "true").lower() == "true"
    if record_export:
        new_export = Export(
            user_id=current_user.id,
            filename=filename,
            year=year,
            month=month,
            end_year=end_year,
            end_month=end_month
        )
        db.session.add(new_export)
        db.session.commit()

    return send_file(
        buf,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# ─── Import Excel ─────────────────────────────────────────────────────────────
@app.route("/api/import", methods=["POST"])
@login_required
def api_import():
    if not current_user.is_pro:
        return jsonify({"success": False, "error": "Fonctionnalité PRO uniquement."})
    if "file" not in request.files:
        return jsonify({"success": False, "error": "Aucun fichier reçu."})
    f = request.files["file"]
    settings = get_or_create_settings(current_user)
    labels = settings.field_labels

    try:
        wb = openpyxl.load_workbook(f)
        ws = wb.active
        headers = [str(c.value or "").strip().lower() for c in ws[1]]

        label_map = {v.lower(): k for k, v in labels.items()}
        label_map.update(
            {
                "date": "date",
                "catégorie": "category",
                "libellé": "label",
                "observation": "observation",
                "quantité": "quantity",
                "entrée": "income",
                "dépense": "expense",
                "solde cumulé": "balance",
            }
        )
        col_map = {}
        for i, h in enumerate(headers):
            if h in label_map:
                col_map[label_map[h]] = i

        added = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            inc = float(row[col_map["income"]] or 0) if "income" in col_map else 0
            exp = float(row[col_map["expense"]] or 0) if "expense" in col_map else 0
            if inc == 0 and exp == 0:
                continue
            tx_type = "income" if inc > 0 else "expense"
            amount = inc if tx_type == "income" else exp
            raw_date = row[col_map["date"]] if "date" in col_map else None
            if isinstance(raw_date, datetime):
                tx_date = raw_date.date()
            elif isinstance(raw_date, date):
                tx_date = raw_date
            elif raw_date:
                try:
                    from dateutil import parser as dp

                    tx_date = dp.parse(str(raw_date)).date()
                except:
                    tx_date = date(
                        request.args.get("year", date.today().year, type=int),
                        request.args.get("month", date.today().month, type=int),
                        1,
                    )
            else:
                tx_date = date(
                    request.args.get("year", date.today().year, type=int),
                    request.args.get("month", date.today().month, type=int),
                    1,
                )
            t = Transaction(
                id=str(uuid.uuid4()),
                user_id=current_user.id,
                type=tx_type,
                amount=amount,
                quantity=int(row[col_map["quantity"]] or 1)
                if "quantity" in col_map
                else 1,
                label=str(row[col_map["label"]] or "") if "label" in col_map else "",
                category=str(row[col_map["category"]] or "Import")
                if "category" in col_map
                else "Import",
                observation=str(row[col_map["observation"]] or "")
                if "observation" in col_map
                else "",
                date=tx_date,
            )
            db.session.add(t)
            added += 1
        db.session.commit()
        return jsonify(
            {
                "success": True,
                "message": f"{added} transaction(s) importée(s) avec succès.",
            }
        )
    except Exception as e:
        db.session.rollback()
        return jsonify(
            {"success": False, "error": f"Erreur lors de l'importation : {str(e)}"}
        )


# ─── Reset Data ───────────────────────────────────────────────────────────────
@app.route("/api/reset", methods=["POST"])
@login_required
def api_reset():
    Transaction.query.filter_by(user_id=current_user.id).delete()
    s = get_or_create_settings(current_user)
    s.currency = "FCFA"
    s.categories = PREDEFINED_CATEGORIES
    s.custom_fields = []
    s.field_order = DEFAULT_FIELD_ORDER
    s.hidden_fields = DEFAULT_HIDDEN_FIELDS
    s.field_labels = DEFAULT_FIELD_LABELS
    db.session.commit()
    return jsonify({"success": True})


# ─── User status API ──────────────────────────────────────────────────────────
@app.route("/manifest.json")
def serve_manifest():
    return send_file("static/manifest.json")

@app.route("/sw.js")
def serve_sw():
    return send_file("static/sw.js")

@app.route("/api/user/me")
@login_required
def api_user_me():
    s = get_or_create_settings(current_user)
    return jsonify(
        {
            "id": current_user.id,
            "first_name": current_user.first_name,
            "company_name": current_user.company_name,
            "whatsapp": current_user.whatsapp,
            "email": current_user.email,
            "is_pro": current_user.is_pro,
            "plan": current_user.plan,
            "settings": s.to_dict(),
        }
    )


# ─── Admin Routes ─────────────────────────────────────────────────────────────
@app.route("/checkmode")
def checkmode():
    gs = GlobalSettings.query.first()
    if not gs:
        gs = GlobalSettings(whatsapp_number="", admin_key="")
        db.session.add(gs)
        db.session.commit()

    # Simple security check
    key = request.args.get("key")
    if gs.admin_key and key != gs.admin_key:
        abort(403)

    users = []
    for u in User.query.all():
        user_dict = {
            "id": u.id,
            "first_name": u.first_name,
            "company_name": u.company_name,
            "whatsapp": u.whatsapp,
            "email": u.email,
            "is_pro": u.is_pro,
            "plan": u.plan,
            "activated_at": u.activated_at,
            "created_at": u.created_at,
            "end_date": None,
        }
        if u.is_pro and u.plan and u.activated_at:
            if u.plan == "monthly":
                end = u.activated_at + relativedelta(months=1)
                user_dict["end_date"] = end.strftime("%d/%m/%Y")
            elif u.plan == "quarterly":
                end = u.activated_at + relativedelta(months=3)
                user_dict["end_date"] = end.strftime("%d/%m/%Y")
            elif u.plan in ("yearly", "annual"):
                end = u.activated_at + relativedelta(years=1)
                user_dict["end_date"] = end.strftime("%d/%m/%Y")
        users.append(user_dict)
    pro_count = User.query.filter_by(is_pro=True).count()
    free_count = User.query.filter_by(is_pro=False).count()
    settings = {"whatsapp_number": gs.whatsapp_number, "admin_key": gs.admin_key}
    return render_template(
        "checkmode.html",
        users=users,
        pro_count=pro_count,
        free_count=free_count,
        settings=settings,
        admin_key=gs.admin_key,
    )


@app.route("/api/admin/users/set-status", methods=["POST"])
def api_admin_set_status():
    gs = GlobalSettings.query.first()
    if not gs:
        gs = GlobalSettings(whatsapp_number="", admin_key="")
        db.session.add(gs)
        db.session.commit()
    data = request.get_json()

    # Security check
    auth_key = data.get("auth_key")
    if gs.admin_key and auth_key != gs.admin_key:
        abort(403)

    user_id = data.get("userId")
    is_pro = data.get("isPro")
    plan = data.get("plan")
    user = User.query.get(user_id)
    if not user:
        return jsonify({"success": False, "error": "User not found"}), 404
    user.is_pro = is_pro
    if is_pro:
        user.plan = plan
        user.activated_at = datetime.now(timezone.utc)
    else:
        user.plan = "free"
    db.session.commit()
    return jsonify({"success": True})


@app.route("/api/admin/settings", methods=["POST"])
def api_admin_settings():
    gs = GlobalSettings.query.first()
    if not gs:
        gs = GlobalSettings(whatsapp_number="", admin_key="")
        db.session.add(gs)
        db.session.commit()
    data = request.get_json()

    # Security check
    auth_key = data.get("auth_key")
    if gs.admin_key and auth_key != gs.admin_key:
        abort(403)

    gs.whatsapp_number = data.get("whatsapp_number", gs.whatsapp_number)
    gs.admin_key = data.get("admin_key", gs.admin_key)
    db.session.commit()
    return jsonify({"success": True})


# ─── Init DB & Run ────────────────────────────────────────────────────────────
with app.app_context():
    db.create_all()

if __name__ == "__main__":
    app.run(debug=True, host="0.0.0.0", port=5000)
